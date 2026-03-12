"""Tests for the XLSX election calendar preprocessor."""

import json
from pathlib import Path

from voter_api.lib.election_calendar.parser import parse_calendar_jsonl
from voter_api.lib.election_calendar.xlsx_preprocessor import preprocess_xlsx_calendar


class TestPreprocessXlsxCalendar:
    """Tests for preprocess_xlsx_calendar."""

    def _create_test_xlsx(self, path: Path) -> Path:
        """Create a minimal test XLSX with election calendar data.

        Args:
            path: Directory to write the file in.

        Returns:
            Path to the created XLSX file.
        """
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None

        # Header row
        ws.append(["ELECTION", "ELECTION DATE", "ADVANCE VOTING DATES", "REGISTRATION DEADLINE"])

        # Data rows
        ws.append(
            [
                "General Primary Election/Nonpartisan Election",
                "05/19/2026",
                "04/27/26 - 05/15/26",
                "04/20/2026",
            ]
        )
        ws.append(
            [
                "General Election/Special Election",
                "11/03/2026",
                "10/13/26 - 10/30/26",
                "10/05/2026",
            ]
        )

        xlsx_path = path / "calendar.xlsx"
        wb.save(xlsx_path)
        return xlsx_path

    def test_produces_valid_jsonl(self, tmp_path: Path) -> None:
        """XLSX with valid data produces parseable JSONL output."""
        xlsx_path = self._create_test_xlsx(tmp_path)
        output_path = tmp_path / "output.jsonl"

        count = preprocess_xlsx_calendar(xlsx_path, output_path)

        assert count == 2
        assert output_path.exists()

        entries = parse_calendar_jsonl(output_path)
        assert len(entries) == 2

    def test_election_names_extracted(self, tmp_path: Path) -> None:
        """Election names are correctly extracted from the XLSX."""
        xlsx_path = self._create_test_xlsx(tmp_path)
        output_path = tmp_path / "output.jsonl"

        preprocess_xlsx_calendar(xlsx_path, output_path)
        entries = parse_calendar_jsonl(output_path)

        names = [e.election_name for e in entries]
        assert "General Primary Election/Nonpartisan Election" in names
        assert "General Election/Special Election" in names

    def test_dates_parsed_correctly(self, tmp_path: Path) -> None:
        """Election dates and registration deadlines are parsed."""
        xlsx_path = self._create_test_xlsx(tmp_path)
        output_path = tmp_path / "output.jsonl"

        preprocess_xlsx_calendar(xlsx_path, output_path)
        entries = parse_calendar_jsonl(output_path)

        primary = next(e for e in entries if "Primary" in e.election_name)
        assert str(primary.election_date) == "2026-05-19"
        assert str(primary.registration_deadline) == "2026-04-20"
        assert str(primary.early_voting_start) == "2026-04-27"
        assert str(primary.early_voting_end) == "2026-05-15"

    def test_empty_workbook_returns_zero(self, tmp_path: Path) -> None:
        """Empty workbook (no header row) returns 0 entries."""
        import openpyxl

        wb = openpyxl.Workbook()
        xlsx_path = tmp_path / "empty.xlsx"
        wb.save(xlsx_path)
        output_path = tmp_path / "output.jsonl"

        count = preprocess_xlsx_calendar(xlsx_path, output_path)

        assert count == 0

    def test_qualifying_period_extracted(self, tmp_path: Path) -> None:
        """Qualifying period dates are extracted from rows above the header."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None

        # Qualifying period row above header
        ws.append(["QUALIFYING PERIOD FOR ALL CANDIDATES", "MARCH 2, 2026 - MARCH 6, 2026"])
        ws.append([])  # blank row
        ws.append(["ELECTION", "ELECTION DATE", "ADVANCE VOTING DATES", "REGISTRATION DEADLINE"])
        ws.append(
            [
                "General Primary",
                "05/19/2026",
                "04/27/26 - 05/15/26",
                "04/20/2026",
            ]
        )

        xlsx_path = tmp_path / "with_qualifying.xlsx"
        wb.save(xlsx_path)
        output_path = tmp_path / "output.jsonl"

        preprocess_xlsx_calendar(xlsx_path, output_path)

        lines = output_path.read_text().strip().splitlines()
        record = json.loads(lines[0])
        assert "qualifying_start" in record
        assert "qualifying_end" in record

    def test_jsonl_output_is_valid_json_lines(self, tmp_path: Path) -> None:
        """Each line of the output is valid JSON."""
        xlsx_path = self._create_test_xlsx(tmp_path)
        output_path = tmp_path / "output.jsonl"

        preprocess_xlsx_calendar(xlsx_path, output_path)

        lines = output_path.read_text().strip().splitlines()
        for line in lines:
            record = json.loads(line)
            assert "election_name" in record
            assert "election_date" in record
