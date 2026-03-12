"""Preprocessor for GA SoS XLSX election calendar files.

Extracts election calendar data from the structured XLSX spreadsheets
published by the Georgia Secretary of State Elections Division and writes
standardized JSONL output.
"""

import json
import re
from datetime import date
from pathlib import Path

from loguru import logger


def _parse_mmddyyyy(value: str) -> date | None:
    """Parse a date in MM/DD/YYYY format.

    Args:
        value: Date string like ``"05/19/2026"``.

    Returns:
        Parsed date or None if parsing fails.
    """
    value = value.strip()
    match = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", value)
    if not match:
        return None
    month, day, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
    return date(year, month, day)


def _parse_mmddyy(value: str) -> date | None:
    """Parse a date in MM/DD/YY format, assuming 2000s century.

    Args:
        value: Date string like ``"04/27/26"``.

    Returns:
        Parsed date or None if parsing fails.
    """
    value = value.strip()
    match = re.match(r"(\d{1,2})/(\d{1,2})/(\d{2})", value)
    if not match:
        return None
    month, day, year_2 = int(match.group(1)), int(match.group(2)), int(match.group(3))
    year = 2000 + year_2
    return date(year, month, day)


def _parse_date_flex(value: str) -> date | None:
    """Parse a date in either MM/DD/YYYY or MM/DD/YY format.

    Args:
        value: Date string.

    Returns:
        Parsed date or None.
    """
    result = _parse_mmddyyyy(value)
    if result is not None:
        return result
    return _parse_mmddyy(value)


def _parse_advance_voting_range(text: str) -> tuple[date | None, date | None]:
    """Extract start and end dates from an advance voting date range string.

    Handles formats like:
    - ``"04/27/26 - 05/15/26"``
    - ``"As soon as possible, but no later than 06/08/26 - 06/12/26"``

    Args:
        text: The advance voting dates text from the calendar.

    Returns:
        Tuple of (start_date, end_date).
    """
    # Find all date-like patterns (MM/DD/YY or MM/DD/YYYY)
    date_patterns = re.findall(r"\d{1,2}/\d{1,2}/\d{2,4}", text)
    if len(date_patterns) >= 2:
        start = _parse_date_flex(date_patterns[0])
        end = _parse_date_flex(date_patterns[1])
        return start, end
    if len(date_patterns) == 1:
        return _parse_date_flex(date_patterns[0]), None
    return None, None


def _parse_registration_deadline(text: str) -> date | None:
    """Extract the primary registration deadline from the registration column.

    The column may contain both a regular deadline and a federal contest
    deadline (prefixed with ``*``). This returns the primary (non-starred) one.

    Args:
        text: Registration deadline text, possibly multiline.

    Returns:
        The primary registration deadline date.
    """
    lines = text.strip().splitlines()
    for line in lines:
        line = line.strip()
        if line.startswith("*"):
            continue
        parsed = _parse_date_flex(line)
        if parsed is not None:
            return parsed
    # Fallback: try any date in the text
    date_patterns = re.findall(r"\d{1,2}/\d{1,2}/\d{2,4}", text)
    for dp in date_patterns:
        if not text[max(0, text.index(dp) - 1)].startswith("*"):
            parsed = _parse_date_flex(dp)
            if parsed is not None:
                return parsed
    return None


def _find_header_row(ws: object) -> int | None:
    """Find the row index containing the column headers.

    Looks for a row containing ``"ELECTION"`` and ``"ELECTION DATE"``
    (or similar) text in its cells.

    Args:
        ws: The openpyxl worksheet to scan.

    Returns:
        1-based row index of the header row, or None if not found.
    """
    for row_idx in range(1, min(ws.max_row + 1, 30)):
        values = [str(c.value or "").strip().upper() for c in ws[row_idx]]
        joined = " ".join(values)
        if "ELECTION" in joined and "DATE" in joined:
            return row_idx
    return None


def preprocess_xlsx_calendar(
    input_path: Path,
    output_path: Path,
) -> int:
    """Extract election calendar data from a GA SoS XLSX file.

    Reads the first worksheet, locates the header row, and extracts
    election name, date, advance voting dates, and registration deadline
    from each data row. Writes standardized JSONL output.

    Args:
        input_path: Path to the source XLSX file.
        output_path: Path for the JSONL output file.

    Returns:
        Number of calendar entries written.

    Raises:
        FileNotFoundError: If the input file does not exist.
        ValueError: If the worksheet structure cannot be parsed.
    """
    import openpyxl

    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.active
    if ws is None:
        msg = f"No active worksheet in {input_path}"
        raise ValueError(msg)

    header_row = _find_header_row(ws)
    if header_row is None:
        logger.warning("Could not find header row in XLSX file {}", input_path)
        output_path.write_text("")
        return 0

    # Map column indices by header text
    headers: dict[str, int] = {}
    for cell in ws[header_row]:
        val = str(cell.value or "").strip().upper()
        if "ELECTION" in val and "DATE" not in val:
            headers["election"] = cell.column - 1
        elif "ELECTION DATE" in val or ("DATE" in val and "ELECTION" in val):
            headers["date"] = cell.column - 1
        elif "ADVANCE" in val or "VOTING" in val:
            headers["advance_voting"] = cell.column - 1
        elif "REGISTRATION" in val or "DEADLINE" in val:
            headers["registration"] = cell.column - 1

    if "election" not in headers or "date" not in headers:
        logger.warning("Could not identify required columns in XLSX header: {}", headers)
        output_path.write_text("")
        return 0

    # Parse qualifying period from the sheet (look for it above the header)
    qualifying_start = None
    qualifying_end = None
    for row_idx in range(1, header_row):
        for cell in ws[row_idx]:
            val = str(cell.value or "")
            if "QUALIFYING" in val.upper():
                # Look for date range in nearby cells
                for search_cell in ws[row_idx]:
                    search_val = str(search_cell.value or "")
                    date_matches = re.findall(
                        r"(?:MARCH|JANUARY|FEBRUARY|APRIL|MAY|JUNE|JULY|AUGUST|SEPTEMBER|OCTOBER|NOVEMBER|DECEMBER)\s+\d{1,2},?\s+\d{4}",
                        search_val,
                        re.IGNORECASE,
                    )
                    if len(date_matches) >= 2:
                        from datetime import datetime

                        for _fmt in ("%B %d, %Y", "%B %d %Y"):
                            try:
                                qualifying_start = datetime.strptime(date_matches[0].replace(",", ""), "%B %d %Y")  # noqa: DTZ007
                                qualifying_end = datetime.strptime(date_matches[1].replace(",", ""), "%B %d %Y")  # noqa: DTZ007
                                # Set end to end of day
                                qualifying_end = qualifying_end.replace(hour=23, minute=59, second=59)
                                break
                            except ValueError:
                                continue
                    # Also try MM/DD/YYYY format
                    date_patterns_found = re.findall(r"\d{1,2}/\d{1,2}/\d{4}", search_val)
                    if len(date_patterns_found) >= 2 and qualifying_start is None:
                        from datetime import datetime

                        d1 = _parse_mmddyyyy(date_patterns_found[0])
                        d2 = _parse_mmddyyyy(date_patterns_found[1])
                        if d1 and d2:
                            qualifying_start = datetime.combine(d1, datetime.min.time())  # noqa: DTZ007
                            qualifying_end = datetime.combine(  # noqa: DTZ007
                                d2,
                                datetime.max.time().replace(microsecond=0),
                            )

    entries: list[dict] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_cells = list(ws[row_idx])
        # Get election name
        election_name_raw = str(row_cells[headers["election"]].value or "").strip()
        if not election_name_raw:
            continue

        # Get election date
        date_val = row_cells[headers["date"]].value
        if date_val is None:
            continue
        date_str = str(date_val).strip()
        election_date = _parse_date_flex(date_str)
        if election_date is None:
            # Try parsing as a datetime object from openpyxl
            if hasattr(date_val, "date"):
                election_date = date_val.date()
            elif hasattr(date_val, "strftime"):
                election_date = date_val
            else:
                logger.warning("Could not parse election date '{}' for '{}'", date_str, election_name_raw)
                continue

        # Get advance voting dates
        early_start = None
        early_end = None
        if "advance_voting" in headers:
            adv_text = str(row_cells[headers["advance_voting"]].value or "").strip()
            if adv_text:
                early_start, early_end = _parse_advance_voting_range(adv_text)

        # Get registration deadline
        reg_deadline = None
        if "registration" in headers:
            reg_text = str(row_cells[headers["registration"]].value or "").strip()
            if reg_text:
                reg_deadline = _parse_registration_deadline(reg_text)

        entry = {
            "election_name": election_name_raw,
            "election_date": str(election_date),
        }
        if reg_deadline:
            entry["registration_deadline"] = str(reg_deadline)
        if early_start:
            entry["early_voting_start"] = str(early_start)
        if early_end:
            entry["early_voting_end"] = str(early_end)
        if qualifying_start:
            entry["qualifying_start"] = qualifying_start.isoformat()
        if qualifying_end:
            entry["qualifying_end"] = qualifying_end.isoformat()

        entries.append(entry)
        logger.debug("Extracted XLSX entry: {} on {}", election_name_raw, election_date)

    with output_path.open("w", encoding="utf-8") as f:
        for entry in entries:
            f.write(json.dumps(entry) + "\n")

    logger.info("Wrote {} calendar entries from XLSX to {}", len(entries), output_path)
    return len(entries)
